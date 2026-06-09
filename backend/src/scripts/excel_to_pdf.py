import fitz
import sys
import openpyxl

def excel_to_pdf(input_path, output_path):
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"Error loading Excel file: {e}", file=sys.stderr)
        sys.exit(1)
        
    html_content = "<html><body style='font-family: Arial, sans-serif; padding: 25px; color: #1e293b;'>"
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Skip empty sheets
        if sheet.max_row <= 1 and sheet.max_column <= 1 and not sheet.cell(1, 1).value:
            continue
            
        html_content += f"<h2 style='color: #1e3a8a; border-bottom: 2px solid #cbd5e1; padding-bottom: 6px; font-size: 16px; margin-top: 20px;'>{sheet_name}</h2>"
        html_content += "<table style='border-collapse: collapse; font-size: 9px; margin-bottom: 30px; width: 100%; min-width: 400px;'>"
        
        # Read rows
        for row in sheet.iter_rows(values_only=False):
            # Check if row is completely empty
            if all(cell.value is None for cell in row):
                continue
                
            html_content += "<tr>"
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                bg_style = "background-color: #f8fafc;" if cell.row == 1 else ""
                # Bold for top header row
                weight_style = "font-weight: bold; color: #0f172a;" if cell.row == 1 else ""
                
                # Align numeric values to the right
                try:
                    float(val)
                    align_style = "text-align: right;"
                except ValueError:
                    align_style = "text-align: left;"
                    
                html_content += f"<td style='border: 1px solid #cbd5e1; padding: 5px; min-width: 40px; {bg_style} {weight_style} {align_style}'>{val}</td>"
            html_content += "</tr>"
        html_content += "</table>"
        
    html_content += "</body></html>"
    
    try:
        fitz_doc = fitz.open(stream=html_content.encode('utf-8'), filetype='html')
        pdf_bytes = fitz_doc.convert_to_pdf()
        with open(output_path, "wb") as f:
            f.write(pdf_bytes)
        fitz_doc.close()
    except Exception as e:
        print(f"Error generating PDF: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python excel_to_pdf.py <input_path> <output_path>", file=sys.stderr)
        sys.exit(1)
    excel_to_pdf(sys.argv[1], sys.argv[2])
