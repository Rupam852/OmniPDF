import fitz
import sys
import os

def pdf_to_xlsx_gemini(input_path, output_path, gemini_key):
    try:
        import google.generativeai as genai
        import openpyxl
        import json
    except ImportError:
        print("Required python packages not found, using direct conversion.", file=sys.stderr)
        pdf_to_xlsx_direct(input_path, output_path)
        return

    try:
        genai.configure(api_key=gemini_key)
        with open(input_path, "rb") as f:
            pdf_data = f.read()
            
        model = genai.GenerativeModel('gemini-2.5-flash')
        prompt = (
            "Extract all tabular data from this PDF document. For each table found, "
            "return the rows in a structured JSON list format. For example:\n"
            "[\n"
            "  {\"SheetName\": \"Table1\", \"Rows\": [[\"Col1\", \"Col2\"], [\"Val1\", \"Val2\"]]}\n"
            "]\n"
            "Return only the valid JSON array without any markdown formatting wrappers."
        )
        
        response = model.generate_content([
            {
                "mime_type": "application/pdf",
                "data": pdf_data
            },
            prompt
        ])
        
        # Clean up any potential markdown json fences
        text = response.text.strip()
        if text.startswith("```json"):
            text = text[7:]
        if text.endswith("```"):
            text = text[:-3]
        text = text.strip()
        
        tables_data = json.loads(text)
    except Exception as e:
        print(f"Gemini API conversion failed: {e}. Falling back to direct conversion...", file=sys.stderr)
        pdf_to_xlsx_direct(input_path, output_path)
        return

    try:
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        for table_idx, table_info in enumerate(tables_data):
            sheet_name = table_info.get("SheetName", f"Table {table_idx + 1}")
            # Ensure unique sheet name of length <= 31
            sheet_name = sheet_name[:30]
            sheet = wb.create_sheet(title=sheet_name)
            
            rows = table_info.get("Rows", [])
            for r_idx, row_vals in enumerate(rows):
                for c_idx, val in enumerate(row_vals):
                    sheet.cell(row=r_idx + 1, column=c_idx + 1, value=val)
                    
        if not wb.sheetnames:
            wb.create_sheet("Empty Sheet")
            
        wb.save(output_path)
    except Exception as e:
        print(f"Error saving XLSX from Gemini output: {e}. Falling back to direct conversion...", file=sys.stderr)
        pdf_to_xlsx_direct(input_path, output_path)

def pdf_to_xlsx_direct(input_path, output_path):
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        doc = fitz.open(input_path)
        table_count = 0
        
        for p_idx, page in enumerate(doc):
            tables = page.find_tables()
            for t_idx, table in enumerate(tables):
                table_count += 1
                sheet_name = f"Page{p_idx + 1} Table{t_idx + 1}"
                sheet = wb.create_sheet(title=sheet_name[:30])
                
                rows = table.extract()
                for r_idx, row in enumerate(rows):
                    for c_idx, val in enumerate(row):
                        sheet.cell(row=r_idx + 1, column=c_idx + 1, value=val)
                        
        if table_count == 0:
            # Fallback: write text to cells row-by-row
            sheet = wb.create_sheet(title="Extracted Text")
            for p_idx, page in enumerate(doc):
                lines = page.get_text().split("\n")
                for l_idx, line in enumerate(lines):
                    sheet.cell(row=l_idx + 1, column=p_idx + 1, value=line.strip())
                    
        wb.save(output_path)
        doc.close()
    except Exception as e:
        print(f"Direct PDF to XLSX conversion failed: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python pdf_to_excel.py <input_path> <output_path> [gemini_api_key]", file=sys.stderr)
        sys.exit(1)
        
    in_p = sys.argv[1]
    out_p = sys.argv[2]
    key = sys.argv[3] if len(sys.argv) > 3 and sys.argv[3].strip() and sys.argv[3] != "undefined" else ""
    
    if key:
        pdf_to_xlsx_gemini(in_p, out_p, key)
    else:
        pdf_to_xlsx_direct(in_p, out_p)
